import os
from typing import Union, AnyStr, BinaryIO, List, Dict

import openpyxl
from openpyxl.cell import WriteOnlyCell

from .utils import normalize, check_normalized
from .csv import write_csv


def convert_sheet(book, sheet=None):

    data = list()
    sheet = sheet or 0

    if type(sheet) is str:
        sheet = book.get_sheet_by_name(sheet)
    else:
        sheet = book.worksheets[sheet]

    headers = None

    for row in sheet.rows:
        if not headers:
            headers = [c.value for c in row]
            continue
        data.append(dict(zip(headers, (c.value for c in row))))

    return data


def read_xlsx(fp: Union[AnyStr, BinaryIO], sheet=None, **kwargs) -> Union[List[Dict], Dict]:
    book = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    if sheet == '*':
        sheet = book.get_sheet_names()

    if isinstance(sheet, list):
        output = dict()
        for sheet_name in sheet:
            data = convert_sheet(book, sheet=sheet_name)
            output[sheet_name] = data
        return output
    else:
        return convert_sheet(book, sheet=sheet)


def write_xlsx(data: List[Dict], fp: Union[AnyStr, BinaryIO], headers=None, sheet_name=None,
               normalize_data=True, default=''):
    """
    Write a Python object to a XLS or XLSX file based on xlrd.
    :param filename: XLS or XLSX file'
    :param headings: list of headings. If None, will compile form data.
    """

    book = openpyxl.Workbook(write_only=True)

    sheet = book.create_sheet(title=sheet_name)

    if normalize_data:
        data = normalize(data, headers=headers, default=default)
    elif not check_normalized(data, headers=headers):
        raise ValueError(
            'Data must be normalized for XLS files (i.e. all dicts must have same keys). Suggestion: use normalize_data=True')

    headers = headers or list(data[0].keys())

    sheet.append((WriteOnlyCell(sheet, h) for h in headers))
    for row in ((row[k] for k in headers) for row in data):
        sheet.append(row)

    book.save(fp)


def xlsx_to_csv(fp, csv_folder, ignore_dashed_names=True):

    book = openpyxl.load_workbook(fp, read_only=True, data_only=True)

    sheet_names = book.get_sheet_names()

    for sheet_name in sheet_names:
        if ignore_dashed_names and sheet_name.startswith('_'):
            continue
        data = convert_sheet(book, sheet=sheet_name)
        file_name = os.path.join(csv_folder, sheet_name + '.csv')
        write_csv(data, file_name)
