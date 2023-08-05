"""
A script for outputting the results of a PyEHub model.

Outputs can be stored to an excel spreadsheet or printed to the console.
"""

import collections
from collections import OrderedDict
from typing import Union
import pandas as pd
from openpyxl import load_workbook

EXCEL_ENGINE = 'openpyxl'
EXCEL_FIRST_COLUMN_WIDTH = 25

def to_dataframes(frames: dict) -> OrderedDict:
    """
    Convert the values of a dictionary into dataframes if they can be converted.

    Args:
        frames: The dictionary to be converted

    Returns:
        An ordered dictionary with dataframes as values (if they can be
        converted).
    """
    for name, value in frames.items():
        if isinstance(value, dict):
            value = to_dataframe(name, value)

        frames[name] = value

    return OrderedDict(sorted(frames.items()))


def to_dataframe(name: str, value: dict) -> pd.DataFrame:
    """
    Convert a dictionary into a dataframe.

    Args:
        name: The name for the dataframe
        value: The dictionary to be converted

    Returns:
        A Pandas DataFrame.
    """
    value = sort_dict(value)
    value = pd.DataFrame.from_dict(value, orient='index')

    # To remove confusion on what the column '0' means
    if list(value.columns) == [0]:
        value.columns = [name]

    # Make wide matrices fit on the screen
    num_rows, num_columns = value.shape
    if num_columns > num_rows:
        value = value.T

    return value


def _create_sheet(excel_writer: pd.ExcelWriter, sheet_name: str):
    """
    Create an Excel sheet with the given name.

    Args:
        excel_writer: The ExcelWriter object
        sheet_name: The name of the new sheet

    Returns:
        The worksheet for the new sheet
    """
    # Create an empty dataframe just so that we can create a sheet
    pd.DataFrame().to_excel(excel_writer, sheet_name)

    return excel_writer.sheets[sheet_name]


def output_excel(solution: dict, file_name: str, time_steps: int = None, sheets:list = None):
    """ 
    Output the solution dictionary to an Excel file.

    It outputs the time series data in their own sheet with the rest being put
    into another sheet.

    Args:
        solution: A dictionary of the solution part of the response format.
            This contains the variables and parameters of the solved model.
        file_name: The name of the Excel file to write to
        time_steps: Optional. The number of time steps to classify a dataframe
            as holding time series data.
        sheets: A list of all the sheets to be contained in the excel file.
    """
    excel_writer = pd.ExcelWriter(file_name, engine=EXCEL_ENGINE)
    attributes = to_dataframes(solution)

    def _has_time_series_data(value):
        if isinstance(value, pd.DataFrame):
            num_rows, _ = value.shape

            return time_steps and num_rows >= time_steps
        return False

    time_series = {name: value for name, value in attributes.items()
                   if _has_time_series_data(value)}
    non_time_series = {name: value for name, value in attributes.items()
                       if name not in time_series}

    for name, value in time_series.items():
        value.to_excel(excel_writer, f'{name}')

    #: Set first_sheet to "Other" if sheets is not passed, otherwise set it to
    #: the first item in sheets.
    first_sheet = 'Other' if sheets is None else sheets[0]

    #: If sheets is passed, then pass all the sheets except the first as additional sheets.
    if sheets is not None:
        _dict_to_excel_sheet(excel_writer, first_sheet, non_time_series, padding=1, additional_sheets= sheets[1:])
    else:
        #: Otherwise dont sent any value for additional sheets, as none have been provided.
        _dict_to_excel_sheet(excel_writer, first_sheet, non_time_series, padding=1)

    excel_writer.save()


def _dict_to_excel_sheet(excel_writer, sheet_name, non_time_series, padding=0, additional_sheets: list=None):
    """ 
    Output a dictionary to an excel sheet.

    Args:
        excel_writer: The ExcelWriter object
        sheet_name: The name of the sheet to output to
        non_time_series: The dictionary of data
        padding: The padding between key, value pairs
        additional_sheets: a list containing names of other sheets
                           to potentially be added to the excel document.
    """
    hold_row = 1
    row = 1
    hold_sheet_name = sheet_name
    sheet = _create_sheet(excel_writer, sheet_name)
    hold_sheet = sheet

    for name, value in non_time_series.items():
        if additional_sheets is not None and name in additional_sheets:
            hold_row = row
            row = 1
            sheet = _create_sheet(excel_writer, name)
            sheet_name = name
        if isinstance(value, pd.DataFrame):
            num_rows, num_cols = value.shape

            # Compensate for empty Excel row of dataframes. No idea why. Gee thanks
            # Also handle case for when printing first value in sheet.
            startrow = max(row - 1, 1)
            value.to_excel(excel_writer, sheet_name, startrow=startrow)

            if num_cols > 1:
                # Output variable name if not in column header
                sheet.cell(row=row, column=1).value = name

            row += num_rows + 1  # Add row for column headers
        else:
            sheet.cell(row=row, column=1).value = name

            if not isinstance(value, collections.Iterable):
                # Simplify writing column values
                value = [value]

            for col, x in enumerate(value, start=2):
                sheet.cell(row=row, column=col).value = str(x)

            row += 1

        if sheet_name != hold_sheet_name:
            sheet_name = hold_sheet_name
            row = hold_row
            sheet = hold_sheet

        row += padding

    sheet.column_dimensions['A'].width = EXCEL_FIRST_COLUMN_WIDTH


def pretty_print(results: dict) -> None:
    """Print the results in a prettier format.

    Args:
        results: The results dictionary to print
    """
    version = results['version']
    solver = results['solver']

    print("Version: {}".format(version))

    print("Solver")
    print(f"\ttermination condition: {solver['termination_condition']}")
    if solver['termination_condition']!='Infeasible':
        print(f"\ttime: {solver['time']}")

        print("Solution")
        print_section('Stuff', results['solution'])


def print_section(section_name: str, solution_section: dict) -> None:
    """
    Print all the attributes with a heading.

    Args:
        section_name: The heading
        solution_section: The dictionary with all the attributes
    """
    half_heading = '=' * 10
    print(f"\n{half_heading} {section_name} {half_heading}")

    attributes = to_dataframes(solution_section)
    for name, value in attributes.items():
        with pd.option_context('display.max_rows', None, 'display.max_columns', 3):
            print(f"\n{name}: \n{value}")


def sort_dict(mapping: Union[dict, OrderedDict]) -> OrderedDict:
    """
    Sorts a dictionary and all its sub-dicionaries as well.

    Examples:
        >>> sort_dict({1: 'a', 3: 'c', 2: 'b'})
        OrderedDict([(1, 'a'), (2, 'b'), (3, 'c')])
        >>> sort_dict({1: {3: 'c', 2: 'b', 1: 'a'}})
        OrderedDict([(1, OrderedDict([(1, 'a'), (2, 'b'), (3, 'c')]))])

    Args:
        mapping: The dictionary to be sorted

    Returns:
        The sorted dictionary as an OrderedDict
    """
    if not isinstance(mapping, dict):
        return mapping

    for key, value in mapping.items():
        mapping[key] = sort_dict(value)

    return OrderedDict(sorted(mapping.items()))


def print_capacities(results):
    """
    Prints the capacities of each tech and storage at the end

    :param results: the solved model

    """
    solution = results['solution']
    capacity_tech = solution['capacity_tech']
    capacity_storage = solution['capacity_storage']

    half_heading = '=' * 10
    print(f"\n{half_heading} {'Capacities'} {half_heading}")
    print(half_heading * 4)
    print(capacity_storage)
    print(half_heading * 4)
    print(capacity_tech)


def print_warning(results):
    """
    Prints an error if the model burns energy, i.e there is energy from storage and energy to storage
    at same time step.

    :param results: the solved model

    """
    solution = results['solution']
    storages = solution['storages']
    energy_from_storage = solution['energy_from_storage']
    energy_to_storage = solution['energy_to_storage']
    time = len(solution['time'])

    half_heading = '=' * 4

    for storage in storages:
        for t in range(time):
            energy_from = energy_from_storage[storage][t]
            energy_to = energy_to_storage[storage][t]

            if (energy_to != 0) and (energy_from != 0):
                print(f"{half_heading} {'Warning:'} {half_heading} \n{ 'Burning energy for storage: '} {storage} "
                      f"{'; For time step: '} {t} ")


def stream_info(results, output_file):
    """
    New output format with the information separated in different sheets for different streams.

    :param results: the solved model
    :param output_file: the output excel file
    """

    solution = results['solution']
    streams = solution['streams']
    load = solution['LOADS']
    energy_from_storage = solution['energy_from_storage']
    energy_to_storage = solution['energy_to_storage']
    energy_exported = solution['energy_exported']
    energy_imported = solution['energy_imported']
    demands = solution['demands']
    time = len(solution['time'])
    export_streams = solution['export_streams']
    import_streams = solution['import_streams']
    storages = solution['storages']
    storage_level = solution['storage_level']

    n = len(storages)


    with pd.ExcelWriter(output_file, engine=EXCEL_ENGINE) as writer:
        writer.book = load_workbook(output_file)

    keep = ['other', 'TIME_SERIES', 'energy_input', 'capacity_tech', 'capacity_storage']
    for sheet_name in writer.book.sheetnames:
        if sheet_name not in keep:
            del writer.book[sheet_name]

    for stream in streams:

        if stream in demands:
            df = pd.DataFrame({
                stream + " Load": [load[stream][t] for t in range(time)]})
            df.to_excel(writer, sheet_name=stream, startrow=time + 2)

        df1 = pd.DataFrame({
            "energy_from " + storage: [energy_from_storage[storage][t] for t in range(time)] for storage in storages})

        df2 = pd.DataFrame({
            "energy_to " + storage:  [energy_to_storage[storage][t] for t in range(time)] for storage in storages})

        df3 = pd.DataFrame({
            "level " + storage: [storage_level[storage][t] for t in range(time)] for storage in storages})

        df1.to_excel(writer, sheet_name=stream, )
        df2.to_excel(writer, sheet_name=stream, startcol=n + 2)
        df3.to_excel(writer, sheet_name=stream, startcol=2*(n + 2))

        if stream in import_streams:
            df4 = pd.DataFrame({
                stream + " imported": [energy_imported[stream][t] for t in range(time)]})
            df4.to_excel(writer, sheet_name=stream, startrow=time + 2, startcol=4)

        if stream in export_streams:
            df5 = pd.DataFrame({
                stream + " exported": [energy_exported[stream][t] for t in range(time)]})
            df5.to_excel(writer, sheet_name=stream, startrow=time + 2, startcol=7)

        writer.save()
