import random
import logging
import openpyxl
from pathlib import Path
from bin.excel_data_handling.data_fields import ApplicationDataFields
from bin.schema.schema import Schema


# TODO ExcelDataHandler probably shouldn't be considered a "utility" anymore.
#  It's grown very matchmaker-specific

class ExcelDataHandler:

    def __init__(self, data_path):
        """Return cleaned up, ready-to-be-used excel applicant while logging applicant-quality messages to user

        :param data_path: Path object - path to excel file.
        :return: False if there were errors
        """

        self.__workbook_is_validated = False
        self.__path = data_path
        self.__validated_workbook = None

    def __validate_workbook(self, unvalidated_workbook):
        self.__log_unneeded_worksheets(unvalidated_workbook)
        validated_workbook = dict()
        for applicant_group in Schema.applicant_groups:
            if self.__worksheet_exists(unvalidated_workbook, applicant_group):
                unvalidated_worksheet = unvalidated_workbook[applicant_group]
                wb = self.__validate_worksheet(unvalidated_worksheet, applicant_group)
                validated_workbook[applicant_group] = wb
            else:
                return False
        return validated_workbook

    def __validate_worksheet(self, unvalidated_worksheet, applicant_group):
        validated_worksheet = []
        fields = ApplicationDataFields(applicant_group)
        for unvalidated_applicant in unvalidated_worksheet:
            validated_applicant = self.__validate_applicant(unvalidated_applicant, fields.get_fields())
            validated_worksheet.append(validated_applicant)
        fields.log_missing_and_unused_fields(unvalidated_worksheet)
        self.__log_worksheet_summary(validated_worksheet, applicant_group)
        return validated_worksheet

    @staticmethod
    def __log_worksheet_summary(worksheet_, applicant_group):
        applicant_count = len(worksheet_)
        logging.info(f'Counting {applicant_group} ...')
        if applicant_count == 0:
            logging.warning('None found!\n')
        else:
            logging.info(f'Found {applicant_count}. Nice job!')
            string_of_applicants = 'Here are several randomly selected __applicants:\n'
            for i in list(range(3)):
                rand_index = random.randint(0, applicant_count - 1)
                string_of_applicants += f'\n{worksheet_[rand_index]}\n'
            logging.info(string_of_applicants)

    @staticmethod
    def __worksheet_exists(workbook_, worksheet_name):
        logging.info(f'Looking for `{worksheet_name}` worksheet ...')
        found = False
        if worksheet_name in workbook_:
            logging.info('Found. Nice job!\n')
            found = True
        else:
            logging.error(f'Not found! Make sure your workbook contains a `{worksheet_name}`` worksheet\n')
        return found

    @staticmethod
    def __log_unneeded_worksheets(workbook_):
        """Report to user if there are worksheets other than 'mentors' and 'mentees'."""
        logging.info('Checking for unneeded worksheets ...')

        all_workbooks = set(workbook_.keys())
        desired_workbooks = set(Schema.applicant_groups)
        unneeded_workbooks = all_workbooks - desired_workbooks

        if len(unneeded_workbooks) == 0:
            logging.info('None found. Nice job!\n')
        else:
            logging.warning(f'Found some unneeded workbooks: {unneeded_workbooks}\n')

    @staticmethod
    def __validate_applicant(unvalidated_applicant, fields):
        validated_applicant = dict()
        for field in fields:
            key = field.get_name()
            value = field.validate_field(unvalidated_applicant)
            validated_applicant[key] = value
        return validated_applicant

    def generate_validated_workbook_(self):
        if not self.__workbook_is_validated:
            unvalidated_workbook = self.__get_dicts_from_excel(self.__path)
            self.__validated_workbook = self.__validate_workbook(unvalidated_workbook)
            self.__workbook_is_validated = True
        return self.__validated_workbook
        # TODO this should return None is there's an error, not False.

    @staticmethod
    def __get_dicts_from_excel(path):
        """Get excel applicant. Log status.
        :param path: pathlib.Path() object

        """
        # Return a dictionary. Keys = worksheet names; Values = worksheet contents
        #         Those Values are themselves lists of dictionaries. Each list item is a worksheet row.
        #         Each worksheet row is a dictionary. Keys = column names; Values = cell values

        logging.info(f'Looking for excel applicant in `{path}` ...')
        try:
            wb = openpyxl.load_workbook(path)
            logging.info('Found!\n')
        except:
            logging.error('Not found!\n')
            return False

        workbook_contents = dict()
        for worksheet_name in wb.sheetnames:
            ws = wb[worksheet_name]
            worksheet_contents = []
            # Convert each row to a dictionary, using the column headers as keys
            for row in range(2, ws.max_row + 1):
                row_contents = dict()
                for col in range(1, ws.max_column + 1):
                    key = ws.cell(1, col).value
                    value = ws.cell(row, col).value
                    # logging.debug(value)
                    row_contents[key] = value
                worksheet_contents.append(row_contents)
            workbook_contents[worksheet_name] = worksheet_contents

        return workbook_contents


if __name__ == '__main__':
    # Set up logging
    logging.basicConfig(format='%(levelname)s: %(message)s', level=logging.INFO)

    # Function test
    excel_path = Path(__file__).parent.parent.parent / 'applicant' / 'private' / 'applications.xlsx'
    excel_data_handler = ExcelDataHandler(excel_path)
    excel_data = excel_data_handler.generate_validated_workbook_()

    # if excel_data:
    #     for mentor in excel_data['mentors']:
    #         print(mentor)
    # else:
    #     logging.error('Oh no!!')
