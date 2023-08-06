import openpyxl
import pathlib


__version__ = '0.1.0'


class Scrapper():
    def __init__(self):
        self.result = []
        self._fields = []


    def field(self, name):
        field = Field(name)
        self._fields.append(field)
    

    def table(self, name, fields= []):
        pass


    def scrap(self, targetFileName):
        targets = pathlib.Path().glob(targetFileName)
        #print('targets', list(targets))
        for target in targets:
            self.scrap_one(target)
    

    def scrap_one(self, fileName):
        wb = openpyxl.load_workbook(fileName) 
        #wb = xlrd.open_workbook(excel_file_name)
        ws = wb.worksheets[0]
        res = {}
        for field in self._fields:
            res[field.name] = field.scrap(ws)
        self.result.append(res)
        return res


class Field():
    def __init__(self, name: str):
        self.name = name
        self.cache = None


    def find_value(self, ws, value):
        #print('value not in cache')
        for row in range(ws.min_row, ws.max_row):
            for col in range(ws.min_column, ws.max_column):
                #print(ws.cell(row, col))
                if value == ws.cell(row, col).value:
                    return (row, col)
        return (None, None)


    def find_value_cache(self, ws, value):
        if self.cache is None:
            self.cache = self.find_value(ws, value)
        return self.cache


    def scrap(self, ws):
        row, col = self.find_value_cache(ws, self.name)
        if row is None:
            return None
        return ws.cell(row, col + 1).value

        res = ''
        return res