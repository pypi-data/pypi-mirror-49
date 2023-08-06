from openpyxl import load_workbook
from pathlib import Path


def test_openpyxl():
    path = Path(__file__).parent / "excel.xlsx"
    wb = load_workbook(
        path, read_only=True, data_only=True, keep_links=True
    )
    ws = wb["main"]
    print(ws)
    values = []
    for row in range(1, ws.max_row + 1):
        values.append(ws.cell(row, 1).value)
        print(row)
    values.append(int('5'))
    for value in values:
        print(value, type(value))

test_openpyxl()
