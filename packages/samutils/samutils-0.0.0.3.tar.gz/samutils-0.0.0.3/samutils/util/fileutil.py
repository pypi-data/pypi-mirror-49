#############################
###    对文件的基本操作    ###
#############################
import csv
import datetime
import os
import random
from abc import abstractmethod
from typing import Callable, List

import openpyxl
import xlrd
from openpyxl import Workbook
from samutils import export_dir_path, import_dir_path, export_csv_path

from samutils.util.assisttool import list2dict, formatter_datetime

from samutils.util.logutil import get_default_logger, LoggerUtil
from xlrd import xldate_as_tuple


default_logger = get_default_logger()


def assert_is_expect_file_type(file_path: str, *expect_file_type_list: str):
    """
    判断文件是否属于期望的文件类型
    """
    for file_type in expect_file_type_list:
        if file_path.endswith(file_type):
            return True
    return False


def get_file_name_from_path(file_path):
    """
    从文件路径中获取文件名 如果没有找到文件名,则返回时间戳
    """
    find_file_path = os.path.basename(file_path)
    if find_file_path:
        return find_file_path
    else:
        return datetime.datetime.now().isoformat()


def get_file_path_list_by_dir_path(dir_path):
    """
    从文件目录中获取文件列表
    """
    file_path_list = []
    ls = os.listdir(dir_path)
    for i in range(len(ls)):
        path = os.path.join(dir_path, ls[i])
        file_path_list.append(path)
    return file_path_list


def delete_files_by_dir_path(dir_path: str):
    """ 根据文件夹路径删除文件目录下面的文件 """
    # 列出目录下的文件
    files = os.listdir(dir_path)
    for file in files:
        file_path = f"{dir_path}/{file}"
        if os.path.exists(file_path):
            # 删除文件
            os.remove(file_path)


def get_now_str():
    now = datetime.datetime.now()
    return f"{now.year}{now.month}{now.day}{now.hour}{now.minute}{now.second}{now.microsecond}"


def setup_export_file_path(export_dir_base_path: str = export_dir_path
                           , export_file_name: str = "export"
                           , prefix=None
                           , suffix=None
                           , file_type: str = "csv"
                           , is_batch: bool = False
                           ) -> str:
    """ 设置导出文件路径 """
    file_path_element_list = [f"{export_dir_base_path}{os.sep}{export_file_name}"]
    has_prefix_suffix = True
    if prefix is not None or suffix is not None:
        file_path_element_list.append("(")
        if prefix is not None:
            file_path_element_list.append(f"{prefix}")
        else:
            has_prefix_suffix = False
        if suffix is not None:
            file_path_element_list.append(f"{suffix}")
        else:
            has_prefix_suffix = False
        if has_prefix_suffix:
            file_path_element_list.insert(3, "_")
        file_path_element_list.append(")")
    if is_batch:
        file_path_element_list.append(f"_{get_now_str()}.{file_type}")
    else:
        file_path_element_list.append(f".{file_type}")
    return "".join(file_path_element_list)


def get_csv_export_file_path(export_file_name: str = "export") -> str:
    return f"{export_csv_dir_path}{os.sep}{export_file_name}.csv"


#############################
###   对csv文件的基本操作  ###
#############################


def _assert_single_row(d):
    if not d:
        raise RuntimeError("传入的参数不能为 空")
    if isinstance(d, list):
        dd = []
        dd.extend(d)
        while len(dd):
            pop = dd.pop()
            if pop:
                if type(pop) == list:
                    # print("两层嵌套")
                    return False
                else:
                    # print("单层嵌套")
                    return True
            else:
                # 继续循环
                pass

    else:
        s = f"错误的类型：{type(d)}. 只能处理 list"
        raise RuntimeError(s)


def _csv_write(target_file, mode, write_data, column: list = None):
    if not assert_is_expect_file_type(target_file, ".csv"):
        raise RuntimeError(f"文件 {target_file} 不是 有效的 csv 文件")
    if write_data:
        # 注意这里如果以‘w’的形式打开，每次写入的数据中间就会多一个空行，所以要用‘wb’
        with open(target_file, mode, newline='', encoding="utf-8") as f:
            writer = csv.writer(f, dialect='excel')
            if column:
                writer.writerow(column)
            if not _assert_single_row(write_data):
                # 这里如果需要写入多行，那么就采用循环进行循环输入就可以啦哈
                for line in write_data:
                    if line:
                        writer.writerow(line)
            else:
                writer.writerow(write_data)


def csv_write(target_file: str, write_data: list, write_type: str = "append", column: list = None):
    if write_type not in ["cover", "append"]:
        msg = f"仅支持 cover 和  append 模式， 不支持当前模式: {write_type} "
        raise RuntimeWarning(msg)

    if "cover" == write_type:
        csv_cover_write(target_file, write_data, column=column)
    elif "append" == write_type:
        csv_append_write(target_file, write_data, column=column)


def csv_cover_write(target_file: str, write_data: list, column: list = None):
    """
    覆盖的方式 写入文件
    :param column:
    :param target_file:
    :param write_data:
    :return:
    """
    _csv_write(target_file, 'w', write_data, column=column)


def csv_append_write(target_file: str, write_data: list, column: list = None):
    """
    追加的方式 写入文件
    :param column:
    :param target_file:
    :param write_data:
    :return:
    """
    _csv_write(target_file, 'a+', write_data, column=column)


def csv_read(target_file
             , column: List[str] = None
             , option: str = None
             , start: int = None
             , end: int = None
             , datetime_formatter: str = None
             , correct_func: Callable[[List[dict]], object] = None
             ):
    """
    读csv文件的内容 返回 list
    :param correct_func:
    :param datetime_formatter:
    :param end:
    :param start:
    :param option:
    :param column:
    :param target_file:
    :return:
    """
    if not assert_is_expect_file_type(target_file, "csv"):
        raise RuntimeError(f"文件 {target_file} 不是 有效的 csv 文件")
    if not option:
        option = "list"

    with open(target_file, "r", encoding="utf-8") as f:
        lines = csv.reader(f)
        new_lines = []
        if not start:
            start = 0
        if end:
            for num, line in enumerate(lines):
                if start <= num <= end:
                    if "list" == option:
                        new_lines.append(line)
                    elif "dict" == option and column:
                        res_dict = list2dict(column, line)
                        if correct_func:
                            correct_func(res_dict)
                        new_lines.append(res_dict)
                    else:
                        error_msg = f"不支持当前操作数 {option} 或 操作数为 dict 时, column 为 空"
                        raise RuntimeWarning(error_msg)
        else:
            for num, line in enumerate(lines):
                if num >= start:
                    new_line = formatter_datetime(line, datetime_formatter)
                    if "list" == option:
                        new_lines.append(new_line)
                    elif "dict" == option and column:
                        res_dict = list2dict(column, new_line)
                        if correct_func:
                            correct_func(res_dict)
                        new_lines.append(res_dict)
                    else:
                        error_msg = f"不支持当前操作数 {option} 或 操作数为 dict 时, column 为 空"
                        raise RuntimeWarning(error_msg)

        return new_lines


#############################
###   对xlsx文件的基本操作 ###
#############################


def excel_read_single_sheet(file_path, sheet_index=0, is_skip_first_line=True):
    if not assert_is_expect_file_type(file_path, ".xlsx", ".xls"):
        raise RuntimeError(f"文件 {file_path} 不是 有效的 excel 文件")
    # 文件位置
    excel_file = xlrd.open_workbook(file_path)
    sheet = excel_file.sheet_by_index(sheet_index)
    rows = sheet.nrows
    cols = sheet.ncols
    all_content = []
    for i in range(rows):
        # 跳过第一行
        if is_skip_first_line:
            if i == 0:
                continue
        row_content = []
        for j in range(cols):
            # 表格的数据类型
            cell_type = sheet.cell(i, j).ctype
            cell = sheet.cell_value(i, j)
            if cell_type == 2 and cell % 1 == 0:  # 如果是整形
                cell = int(cell)
            elif cell_type == 3:
                # 转成datetime对象
                if cell:
                    date = datetime(*xldate_as_tuple(cell, 0))
                    cell = date.strftime('%Y-%m-%d %H:%M:%S')
            elif cell_type == 4:
                cell = True if cell == 1 else False
            row_content.append(cell)
        all_content.append(row_content)
    return all_content


def excel_write_single_sheet(write_file_path: str
                             , write_item_list: list
                             , column: list = None
                             ):
    if not assert_is_expect_file_type(write_file_path, ".xlsx", ".xls"):
        raise RuntimeError(f"文件 {write_file_path} 不是 有效的 excel 文件")
    workbook = Workbook()
    worksheet = workbook.active
    if column:
        worksheet.append(column)
    for line in write_item_list:
        worksheet.append(line)
    workbook.save(write_file_path)


def excel_write_single_sheet_append(write_file_path: str
                                    , append_item_list: list):
    new_workbook = openpyxl.load_workbook(write_file_path)
    new_worksheet = new_workbook.active
    for append_item_line in append_item_list:
        new_worksheet.append(append_item_line)

    new_workbook.save(write_file_path)


def convert_xls_2_xlsx(original_file_path) -> str:
    if not assert_is_expect_file_type(original_file_path, ".xls"):
        raise RuntimeError(f"文件 {original_file_path} 不是 有效的 xls 文件")
    new_file_path = f"{original_file_path}x"
    if os.path.exists(original_file_path):
        content = excel_read_single_sheet(original_file_path)
        excel_write(new_file_path, content)
        if os.path.exists(original_file_path):
            os.remove(original_file_path)
    return new_file_path


def excel_write(write_file_path: str
                , write_item_list: list
                , column: list = None
                , write_type: str = "cover"
                , auto_convert_xls_2_xlsx: bool = True):
    if not assert_is_expect_file_type(write_file_path, ".xlsx"):
        if auto_convert_xls_2_xlsx:
            write_file_path = convert_xls_2_xlsx(write_file_path)
        else:
            raise RuntimeError(f"文件 {write_file_path} 不是 有效的 xlsx 文件, 需要转换成为 xlsx 文件才能操作")
    if "cover" == write_type:
        if os.path.exists(write_file_path):
            os.remove(write_file_path)
        excel_write_single_sheet(write_file_path, write_item_list, column=column)
    else:
        if os.path.exists(write_file_path):
            excel_write_single_sheet_append(write_file_path, write_item_list)
        else:
            excel_write_single_sheet(write_file_path, write_item_list, column=column)


#############################
###   对通用文件的基本操作 ###
#############################


def read_file(file_path: str
              , is_skip_first_line: bool = False
              , result_type: str = "list"
              , column: list = None
              , result_formatter: Callable[[list, list], list] = None
              , start: int = 0
              , end: int = None
              , max_random_count: int = None
              ):
    """
    读取文件:
    1: 支持的文件格式有[ .csv .xls .xlsx .log .txt]
    2: 支持对读取的结果格式化
    3: 支持随机读取，主要用于判断文件格式的是否一致
    4: 支持跳过第一行, 默认是不跳过
    5: 支持读取指定的起止区间
    """
    if result_type not in ["list", "dict"]:
        raise RuntimeError(f"不支持 当前的可选项 {result_type}, 仅支持 list, dict 以及自定义的 formatter")
    if result_type == "dict":
        if not column:
            raise RuntimeError(f"当你期望读取结果格式化为字典时,请务必填写 column 参数")

    if file_path.endswith(".csv") or file_path.endswith(".xlsx") or file_path.endswith(".xls") or file_path.endswith(
            ".log") or file_path.endswith(".txt"):
        if file_path.endswith(".csv"):
            if is_skip_first_line:
                read_result = csv_read(file_path)[1:]
            else:
                read_result = csv_read(file_path)
        elif file_path.endswith(".log") or file_path.endswith(".txt"):
            with open(file_path, "r", encoding="utf-8") as f:
                read_result = f.readlines()
        else:
            read_result = excel_read_single_sheet(file_path, is_skip_first_line=is_skip_first_line)

        end = end if end else len(read_result)
        end = end if len(read_result) >= end else len(read_result)
        need_handle_result = read_result[start: end]
        random_read_result = []
        if max_random_count and max_random_count > 0:
            max_random_count = len(need_handle_result) if len(
                need_handle_result) <= max_random_count else max_random_count
            for i in range(max_random_count):
                if i == 0:
                    random_read_result.append(need_handle_result[0])
                else:
                    random_index = random.randint(1, max_random_count - 1)
                    random_read_result.append(need_handle_result[random_index])

        need_handle_result = random_read_result if random_read_result else need_handle_result
        if result_formatter:
            # 如果 自定义了 格式化方法 就不走 系统预设的结果格式化方法
            return result_formatter(need_handle_result, column)
        else:
            if "list" == result_type:
                return need_handle_result
            else:
                new_read_result = []
                for res in need_handle_result:
                    new_read_result.append(list2dict(column, res))
                return new_read_result

    else:
        raise RuntimeError(f"不支持 对 当前文件 {file_path} 的 读写 ")


def write_file(file_path: str, data_list: list, column: list = None, write_type="cover"):
    if file_path.endswith(".csv"):
        csv_write(file_path, data_list, column=column, write_type=write_type)
    elif file_path.endswith(".xlsx") or file_path.endswith(".xls"):
        excel_write(file_path, data_list, column=column, write_type=write_type)
    else:
        msg = f"不支持 对 当前文件 {file_path} 的 读写 "
        raise RuntimeError(msg)


def write_file_quick(data_list: list
                     , export_file_name: str = "quick_export"
                     , column: list = None
                     , prefix: str = None
                     , suffix: str = None
                     , file_type: str = "csv"
                     , write_type="cover"):
    export_file_path = setup_export_file_path(export_file_name=export_file_name, prefix=prefix, suffix=suffix,
                                              file_type=file_type)
    if export_file_path.endswith(".csv"):
        csv_write(export_file_path, data_list, column=column, write_type=write_type)
    elif export_file_path.endswith(".xlsx") or export_file_path.endswith(".xls"):
        excel_write(export_file_path, data_list, column=column, write_type=write_type)
    else:
        msg = f"不支持 对 当前文件 {export_file_path} 的 读写 "
        raise RuntimeError(msg)


def check_file(file_path: str, import_table_column_list: list = None):
    if import_table_column_list:
        default_logger.info(f"将对文件: {file_path} 的内容进行检查...")
        lines = read_file(file_path, is_skip_first_line=False)
        if lines:
            column_line_str = ''.join(lines[0])
            table_column_str = ''.join(import_table_column_list)
            if column_line_str.lower() == table_column_str.lower():
                column_length = len(import_table_column_list)
                for line in lines:
                    if line and column_length == len(line):
                        continue
                    else:
                        default_logger.info(f"记录: {line} 长度不够!")
                        default_logger.info(f"注意: 文件: {file_path} 的内容检查不通过")
                        return False
                default_logger.info(f"文件: {file_path} 的内容检查通过")
                return True
            else:
                default_logger.info(f"注意: 文件: {file_path} 的内容检查不通过")
                return False
        else:
            default_logger.info(f"文件:{file_path} 表头与预设定的表头不一致")
            msg = f"文件: {file_path} 读取到的内容为空."
            raise RuntimeError(msg)
    else:
        print(f"注意: 没有对文件: {file_path} 的内容进行检查.")
        return True


class FileUtil(LoggerUtil):
    def __init__(self, base_dir_path: str = import_dir_path, import_table_column_list: list = None):
        super().__init__(name="HandFile", level="info")
        self.base_dir_path = base_dir_path
        self.import_table_column_list = import_table_column_list

    def assert_read_file_content_is_suitable(self, file_path) -> bool:
        return check_file(file_path, self.import_table_column_list)

    @staticmethod
    def setup_export_file_path(file_path
                               , prefix=None
                               , suffix=None
                               , file_type: str = "csv") -> str:
        export_file_name = get_file_name_from_path(file_path)
        if "." in export_file_name:
            export_file_name = export_file_name[:export_file_name.find(".")]
        return setup_export_file_path(export_file_name=export_file_name, prefix=prefix, suffix=suffix,
                                      file_type=file_type)

    def _batch_handle(self, handle_single_file_func: Callable[[str], None], start: int = None, end: int = None):
        if not self.base_dir_path:
            raise RuntimeWarning("基本路径不能为空")
        handle_file_path_lit = get_file_path_list_by_dir_path(self.base_dir_path)
        if handle_file_path_lit:
            max_size = len(handle_file_path_lit)
            start_index = start if start else 0
            if not end:
                end_index = 0 if end == 0 else max_size - 1
            else:
                end_index = end
            for i in range(max_size):
                if start_index <= i <= end_index:
                    file_path = handle_file_path_lit[i]
                    self.logger.info(f"开始处理 文件: {file_path}  ...")
                    handle_single_file_func(file_path)
                    self.logger.info(f"处理文件: {file_path} 结束")

        else:
            raise RuntimeWarning("该文件夹是空文件夹")

    def handle_dir(self, start: int = None, end: int = None):
        self._batch_handle(self.handle_file, start=start, end=end)

    @abstractmethod
    def handle_file(self, file_path: str):
        raise NotImplementedError

    @staticmethod
    def file_read(file_path: str
                  , is_skip_first_line: bool = False
                  , result_type: str = "list"
                  , column: list = None
                  , result_formatter: Callable[[list, list], list] = None
                  , start: int = None
                  , end: int = None
                  , max_random_count: int = None
                  ):
        return read_file(
            file_path=file_path
            , is_skip_first_line=is_skip_first_line
            , result_type=result_type
            , column=column
            , result_formatter=result_formatter
            , start=start
            , end=end
            , max_random_count=max_random_count
        )

    @staticmethod
    def file_write(file_path: str, data_list: list, column: list = None, write_type: str = "cover"):
        write_file(file_path=file_path, data_list=data_list, column=column, write_type=write_type)

    def file_write_quick(self, file_path: str = None
                         , data_list: list = None
                         , export_file_name: str = None
                         , column: list = None
                         , prefix=None
                         , suffix=None
                         , write_type: str = "cover"
                         , file_type: str = "csv"
                         ):
        if export_file_name:
            write_file_quick(data_list=data_list
                             , export_file_name=export_file_name
                             , column=column
                             , prefix=prefix
                             , suffix=suffix
                             , write_type=write_type
                             , file_type=file_type)
        else:
            export_file_path = self.setup_export_file_path(file_path, prefix=prefix, suffix=suffix, file_type=file_type)
            write_file(file_path=export_file_path
                       , data_list=data_list
                       , column=column
                       , write_type=write_type
                       )


#############################
###     常见的文件的操作   ###
#############################


class Combine(FileUtil):
    """
    合并文件
    city, term, property
    """

    def __init__(self):
        super().__init__()
        self.export_csv_path = self.setup_export_file_path(export_csv_path, prefix="combine")

    def handle_file(self, file_path: str):
        new_lines = self.file_read(file_path, is_skip_first_line=False)
        self.file_write(self.export_csv_path, new_lines, write_type="append")


def combine_file():
    process = Combine()
    process.handle_dir()


class Csv2Xlsx(FileUtil):
    """
    csv 转换成 excel
    """

    def handle_file(self, file_path: str):
        lines = self.file_read(file_path)
        export_file_path = self.setup_export_file_path(file_path, file_type="xlsx")
        self.file_write(export_file_path, lines)

    def csv2xlsx(self):
        self.handle_dir()


def csv2xlsx():
    process = Csv2Xlsx()
    process.handle_dir()


class Xlsx2Csv(FileUtil):
    """
    xlsx转换成csv
    """

    def handle_file(self, file_path: str):
        lines = self.file_read(file_path)
        export_file_path = self.setup_export_file_path(file_path)
        self.file_write(export_file_path, lines)

    def xlsx2csv(self):
        self.handle_dir()


def xlsx2csv():
    process = Xlsx2Csv()
    process.handle_dir()


class SplitFileProcess(FileUtil):
    """
    分割文件
    """

    def __init__(self, batch_size: int = 10000):
        super().__init__()
        self.batch_size = batch_size

    def handle_file(self, file_path: str, fix_column: bool = True):
        lines = self.file_read(file_path)
        column = lines[0]
        new_lines = []
        file_name = get_file_name_from_path(file_path)
        batch_num = 0
        for i, line in enumerate(lines):
            if i > 0 and i % self.batch_size == 0:
                batch_num += 1
                export_path = setup_export_file_path(export_file_name=file_name, prefix=i + 1 - self.batch_size,
                                                     suffix=i, file_type="xlsx")
                if fix_column:
                    new_lines.insert(0, column)
                self.file_write(export_path, new_lines)
                new_lines.clear()
            new_lines.append(line)
        else:
            export_path = setup_export_file_path(export_file_name=file_name, prefix=batch_num * self.batch_size,
                                                 suffix=len(lines), file_type="xlsx")
            if fix_column:
                new_lines.insert(0, column)
            self.file_write(export_path, new_lines)
            new_lines.clear()

    def split_file(self):
        self.handle_dir()


def split_file():
    process = SplitFileProcess()
    process.split_file()


if __name__ == "__main__":
    f = setup_export_file_path(export_file_name="ok", prefix=0)
    print(f)
    print(get_file_name_from_path("ok/"))
