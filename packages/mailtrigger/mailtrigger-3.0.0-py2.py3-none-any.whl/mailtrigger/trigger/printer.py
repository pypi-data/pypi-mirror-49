# -*- coding: utf-8 -*-

import os
import time

from openpyxl import load_workbook
from openpyxl import Workbook
from .trigger import Trigger, TriggerException
from ..logger.logger import Logger


class Printer(Trigger):
    def __init__(self, config):
        if config is None:
            raise TriggerException('invalid printer configuration')
        self._debug = config.get('debug', False)
        self._filter = config.get('filter', None)
        self._name = config.get('file', 'output.xlsx')

    def _append(self, event):
        def _duplicated(sheet, data):
            if sheet.max_row <= 1:
                return False
            if data['date'] != sheet.cell(row=sheet.max_row, column=2).value:
                return False
            return True
        wb = load_workbook(self._name)
        title = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        if title in wb.sheetnames:
            ws = wb[title]
        else:
            ws = wb.create_sheet(title=title)
            ws.append(['Content', 'Date', 'From', 'Subject', 'To'])
        if _duplicated(ws, event) is False:
            if type(event['to']) is list:
                event['to'] = ','.join(event['to'])
            ws.append([event['content'], event['date'], event['from'], event['subject'], event['to']])
        wb.save(filename=self._name)

    def _check(self, event):
        def _check_helper(data, event):
            if event is None:
                return False
            sender = data.get('from', None)
            if sender is None or event['from'] != sender:
                return False
            subject = data.get('subject', '').strip()
            if event['subject'].startswith(subject) is False:
                return False
            return True
        ret = False
        for item in self._filter:
            if _check_helper(item, event) is True:
                ret = True
                break
        return ret

    def _create(self, event):
        wb = Workbook()
        ws = wb.active
        ws.title = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        ws.append(['Content', 'Date', 'From', 'Subject', 'To'])
        if type(event['to']) is list:
            event['to'] = ','.join(event['to'])
        ws.append([event['content'], event['date'], event['from'], event['subject'], event['to']])
        wb.save(filename=self._name)

    @staticmethod
    def help():
        return ''

    def run(self, event):
        if self._check(event) is False:
            return '', False
        if not os.path.exists(self._name):
            self._create(event)
        else:
            self._append(event)
            Logger.info('appended to %s' % self._name)
        return '', False
